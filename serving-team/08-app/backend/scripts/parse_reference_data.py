#!/usr/bin/env python3
"""기인물 참고자료(KOSHA) 정규화 — 바이너리/마크다운 → 스크립트 사용가능 jsonl.

입력(data-team/05-enrichment/reference-data/기인물참고자료/):
  - SIF 아카이브 xlsx (제조업등 2576 + 건설업 3463): 기인물+재해개요+재해유발요인+위험성감소대책
  - 구조화_전체_사고사례.md (서비스 344) / 구조화_제조업_전체.md (제조 833): ②사고원인 ③예방조치
산출(.../parsed/):
  - sif_archive.jsonl   (1 case/line, 통합 스키마)
  - accident_cases.jsonl(1 case/line, 통합 스키마)

용도: 기인물→절/관→조문 매처를 텍스트 사례 대규모로 검증(라벨 병목 우회), 기인물 alias 사전,
required_measures 보강. (사진 추출 절반은 별도.)

사용: .venv/bin/python scripts/parse_reference_data.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve()
REPO = HERE.parents[4]
REF = REPO / "data-team" / "05-enrichment" / "reference-data" / "기인물참고자료"
OUT = REF / "parsed"
OUT.mkdir(exist_ok=True)


def split_measures(s):
    """▶ 또는 줄바꿈으로 조치 리스트 분리."""
    if not s:
        return []
    parts = re.split(r"[▶\n]+", str(s))
    return [p.strip(" -·\t") for p in parts if p.strip(" -·\t")]


def parse_sif():
    import openpyxl
    p = next(REF.glob("*SIF*xlsx"), None)
    if not p:
        print("[SIF] xlsx 없음"); return 0
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    recs = []
    sheet_domain = {"아카이브(제조업등)": "제조업등", "아카이브(건설업)": "건설업"}
    for sh, domain in sheet_domain.items():
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        if not rows:
            continue
        hdr = [str(c).replace("\n", " ").strip() if c else "" for c in rows[0]]
        idx = {h: i for i, h in enumerate(hdr)}

        def get(row, *names):
            for n in names:
                for h, i in idx.items():
                    if n in h:
                        v = row[i] if i < len(row) else None
                        if v is not None and str(v).strip():
                            return str(v).strip()
            return ""
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            yeonbeon = get(row, "연번")
            gae = get(row, "재해개요")
            if not yeonbeon and not gae:
                continue
            recs.append({
                "source": "SIF", "domain": domain,
                "id": f"SIF-{domain}-{yeonbeon}",
                "gimulmul": get(row, "기인물"),
                "재해개요": gae,
                "고위험상황": get(row, "고위험작업", "고위험"),
                "재해종류": get(row, "재해종류"),
                "재해유발요인": get(row, "재해유발요인"),
                "업종_대분류": get(row, "대분류"),
                "업종_중분류": get(row, "중분류"),
                "업종_소분류": get(row, "소분류"),
                "감소대책": split_measures(get(row, "감소대책")),
            })
    with (OUT / "sif_archive.jsonl").open("w", encoding="utf-8") as f:
        for r in recs:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"[SIF] {len(recs)} → parsed/sif_archive.jsonl")
    return len(recs)


def parse_cases():
    files = {"구조화_전체_사고사례.md": "서비스업", "구조화_제조업_전체.md": "제조업"}
    recs = []
    blk_re = re.compile(r"^##\s+(\d+)\.\s*(.+?)\s*$", re.M)
    for fn, domain in files.items():
        p = REF / fn
        if not p.exists():
            continue
        text = p.read_text(encoding="utf-8")
        marks = list(blk_re.finditer(text))
        for i, m in enumerate(marks):
            num, title = m.group(1), m.group(2)
            body = text[m.end(): marks[i + 1].start() if i + 1 < len(marks) else len(text)]

            def field(label):
                mm = re.search(rf"\*\*{label}\*\*:\s*(.+)", body)
                return mm.group(1).strip() if mm else ""

            def section(marker):
                mm = re.search(rf"###\s*{re.escape(marker)}[^\n]*\n(.+?)(?=\n###|\n---|\Z)", body, re.S)
                if not mm:
                    return []
                lines = [l.strip(" -·\t") for l in mm.group(1).strip().splitlines() if l.strip(" -·\t")]
                return lines

            situ = section("①")
            recs.append({
                "source": "case", "domain": domain, "id": f"case-{domain}-{num}",
                "순번": num, "제목": title,
                "재해유형": field("재해유형"), "재해정도": field("재해정도"),
                "발생일자": field("발생일자"), "발생장소": field("발생장소"),
                "발생상황": " ".join(situ),
                "사고원인": section("②"),
                "예방조치": section("③"),
            })
    with (OUT / "accident_cases.jsonl").open("w", encoding="utf-8") as f:
        for r in recs:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"[cases] {len(recs)} → parsed/accident_cases.jsonl")
    return len(recs)


def main():
    n1 = parse_sif()
    n2 = parse_cases()
    # 샘플 검증 출력
    sif = OUT / "sif_archive.jsonl"
    if sif.exists():
        r = json.loads(sif.read_text(encoding="utf-8").splitlines()[1])
        print("\n[SIF 샘플]", r["id"], "| 기인물:", r["gimulmul"], "| 감소대책:", r["감소대책"][:2])
    cs = OUT / "accident_cases.jsonl"
    if cs.exists():
        r = json.loads(cs.read_text(encoding="utf-8").splitlines()[1])
        print("[case 샘플]", r["id"], "|", r["제목"], "| 예방조치:", r["예방조치"][:1])
    print(f"\n총 {n1 + n2} 사례 정규화 완료 → {OUT}")


if __name__ == "__main__":
    main()
