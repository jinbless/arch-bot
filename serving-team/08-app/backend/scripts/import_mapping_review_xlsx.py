#!/usr/bin/env python3
"""HITL: 작성된 reviewer .xlsx → mapping_review_verdicts UPSERT + 합의/κ 리포트.

export_mapping_review_xlsx.py 산출물(reviewer별)을 읽어 비어있지 않은 판정만 적재(서빙 미반영).
재import = UPSERT(ON CONFLICT(mapping_key, reviewer)) → 멱등. mapping_key별 합의 + Cohen/Fleiss κ
+ disagreements(split) 리포트(json/md/csv).

사용:
  PYTHONIOENCODING=utf-8 python scripts/import_mapping_review_xlsx.py [--input-glob '.../review-xlsx/ohs_mapping_review_*__*.xlsx'] [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[4]
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import text  # noqa: E402

from app.db.database import SessionLocal  # noqa: E402
import _mapping_review_common as mc  # noqa: E402

DEFAULT_GLOB = str(PROJECT_ROOT / "data-team" / "05-enrichment" / "eval-data" / "review-xlsx" / "ohs_mapping_review_*__*.xlsx")


def _reviewer_from_name(p: Path) -> str:
    stem = p.stem
    return stem.rsplit("__", 1)[1] if "__" in stem else stem


def _read_sheet(ws, cols: list) -> list[dict]:
    """헤더는 라벨이라 위치(COLS 순서 + trailing 'reviewer')로 키 매핑."""
    keys = [c[0] for c in cols] + ["reviewer"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {keys[i]: row[i] for i in range(min(len(keys), len(row)))}
        out.append(rec)
    return out


def _verdict_records(path: Path) -> list[dict]:
    reviewer = _reviewer_from_name(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    recs: list[dict] = []
    for sheet, (kind, cols) in mc.SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        for r in _read_sheet(wb[sheet], cols):
            mkey = r.get("mapping_key")
            if not mkey:
                continue
            if sheet == "SCENE":
                payload = {k: r.get(k) for k in ("scene_label_correct", "missing_sr_ids",
                                                 "missing_guide_codes", "recall_complete", "notes")}
                if not any(v not in (None, "") for v in payload.values()):
                    continue
                recs.append({"analysis_id": r.get("analysis_id"), "mapping_kind": "SCENE",
                             "target_id": "scene", "mapping_key": mkey, "row_kind": "scene",
                             "reviewer": reviewer, "verdict": None, "corrected_value": None,
                             "reason_code": None, "notes": json.dumps(payload, ensure_ascii=False),
                             "source_file": path.name})
            else:
                v = (r.get("verdict") or "").strip() if isinstance(r.get("verdict"), str) else r.get("verdict")
                if not v:
                    continue
                target = r.get("sr_id") if kind == "SR" else r.get("guide_code")
                recs.append({"analysis_id": r.get("analysis_id"), "mapping_kind": kind,
                             "target_id": target, "mapping_key": mkey, "row_kind": "candidate",
                             "reviewer": reviewer, "verdict": v,
                             "corrected_value": r.get("corrected_value") or None,
                             "reason_code": r.get("reason_code") or None,
                             "notes": r.get("notes") or None, "source_file": path.name})
    wb.close()
    return recs


_UPSERT = text("""
INSERT INTO mapping_review_verdicts
  (analysis_id, mapping_kind, target_id, mapping_key, row_kind, reviewer, verdict,
   corrected_value, reason_code, notes, source, source_file, updated_at)
VALUES (:analysis_id,:mapping_kind,:target_id,:mapping_key,:row_kind,:reviewer,:verdict,
        :corrected_value,:reason_code,:notes,'excel',:source_file, NOW())
ON CONFLICT (mapping_key, reviewer) DO UPDATE SET
  verdict=EXCLUDED.verdict, corrected_value=EXCLUDED.corrected_value, reason_code=EXCLUDED.reason_code,
  notes=EXCLUDED.notes, row_kind=EXCLUDED.row_kind, source=EXCLUDED.source,
  source_file=EXCLUDED.source_file, updated_at=NOW()
""")


def _cohen_kappa(pairs: list[tuple]) -> float | None:
    """2-rater Cohen κ. pairs=[(a,b)] 라벨쌍. 라벨 1종뿐이면 1.0(완전일치) 처리."""
    n = len(pairs)
    if n == 0:
        return None
    labels = sorted({x for ab in pairs for x in ab})
    po = sum(1 for a, b in pairs if a == b) / n
    if len(labels) <= 1:
        return 1.0
    ca = Counter(a for a, _ in pairs)
    cb = Counter(b for _, b in pairs)
    pe = sum((ca.get(l, 0) / n) * (cb.get(l, 0) / n) for l in labels)
    return round((po - pe) / (1 - pe), 4) if pe < 1 else 1.0


def _agreement(db) -> dict:
    rows = db.execute(text(
        "SELECT mapping_key, mapping_kind, target_id, analysis_id, reviewer, verdict, row_kind "
        "FROM mapping_review_verdicts WHERE row_kind='candidate'")).fetchall()
    by_key: dict[str, dict] = {}
    for mkey, kind, target, aid, reviewer, verdict, _rk in rows:
        d = by_key.setdefault(mkey, {"kind": kind, "target": target, "analysis_id": aid, "verdicts": {}})
        d["verdicts"][reviewer] = verdict
    items, disagreements = [], []
    # per-kind rater label matrices for κ
    for mkey, d in by_key.items():
        vs = d["verdicts"]
        c = Counter(vs.values())
        n_rev = len(vs)
        top, topn = c.most_common(1)[0]
        if n_rev == 1:
            label = "single"
        elif len(c) == 1:
            label = "unanimous_" + top
        elif topn > n_rev / 2:
            label = "majority"
        else:
            label = "split"
        rec = {"mapping_key": mkey, "kind": d["kind"], "target": d["target"],
               "analysis_id": d["analysis_id"], "n_reviewers": n_rev,
               "verdicts": vs, "agreement": label}
        items.append(rec)
        if label == "split":
            disagreements.append(rec)
    # κ: per kind, pairwise over items rated by ≥2
    kappa = {}
    for kind in ("SR", "GUIDE"):
        ks = [it for it in items if it["kind"] == kind and it["n_reviewers"] >= 2]
        reviewers = sorted({r for it in ks for r in it["verdicts"]})
        pair_kappas = []
        for a, b in combinations(reviewers, 2):
            pairs = [(it["verdicts"][a], it["verdicts"][b]) for it in ks
                     if a in it["verdicts"] and b in it["verdicts"]]
            k = _cohen_kappa(pairs)
            if k is not None:
                pair_kappas.append({"raters": f"{a}|{b}", "n": len(pairs), "kappa": k})
        kappa[kind] = {"pairwise": pair_kappas,
                       "mean_kappa": round(sum(p["kappa"] for p in pair_kappas) / len(pair_kappas), 4) if pair_kappas else None,
                       "n_items_multi": len(ks)}
    summary = Counter(it["agreement"] for it in items)
    return {"n_items": len(items), "agreement_breakdown": dict(summary),
            "kappa": kappa, "items": items, "disagreements": disagreements}


def _write_reports(out_dir: Path, date: str, agree: dict):
    base = out_dir / f"mapping_review_agreement_{date}"
    base.with_suffix(".json").write_text(json.dumps(agree, ensure_ascii=False, indent=2), encoding="utf-8")
    # md
    md = [f"# Mapping Review Agreement ({date})", "",
          f"- 검토 pair(candidate): **{agree['n_items']}**",
          f"- 합의 분포: {agree['agreement_breakdown']}", ""]
    for kind, k in agree["kappa"].items():
        md.append(f"- {kind} κ(mean): **{k['mean_kappa']}** (multi-rated {k['n_items_multi']}) — {k['pairwise']}")
    if agree["disagreements"]:
        md += ["", "## 불일치(split) — 재조정 대상", ""]
        for d in agree["disagreements"][:50]:
            md.append(f"- `{d['kind']}` {d['target']} ({d['analysis_id'][:8]}): {d['verdicts']}")
    base.with_suffix(".md").write_text("\n".join(md), encoding="utf-8")
    # csv
    with base.with_suffix(".csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["mapping_key", "kind", "target", "analysis_id", "n_reviewers", "agreement", "verdicts"])
        for it in agree["items"]:
            w.writerow([it["mapping_key"], it["kind"], it["target"], it["analysis_id"],
                        it["n_reviewers"], it["agreement"], json.dumps(it["verdicts"], ensure_ascii=False)])
    return base


def main() -> int:
    import glob
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input-glob", default=DEFAULT_GLOB)
    ap.add_argument("--dry-run", action="store_true", help="DB 미적재, 파싱/합의만")
    args = ap.parse_args()

    files = [Path(p) for p in sorted(glob.glob(args.input_glob))]
    if not files:
        print(f"[warn] 매칭 파일 없음: {args.input_glob}")
        return 1
    print(f"파일 {len(files)}개: {[f.name for f in files]}")

    all_recs = []
    for f in files:
        recs = _verdict_records(f)
        print(f"  {f.name}: 판정 {len(recs)}건 (reviewer={_reviewer_from_name(f)})")
        all_recs.extend(recs)

    db = SessionLocal()
    if not args.dry_run:
        for rec in all_recs:
            db.execute(_UPSERT, rec)
        db.commit()
        print(f"UPSERT {len(all_recs)}건 → mapping_review_verdicts")
    else:
        print(f"[dry-run] UPSERT 생략 ({len(all_recs)}건)")

    agree = _agreement(db) if not args.dry_run else {"n_items": 0, "agreement_breakdown": {}, "kappa": {}, "items": [], "disagreements": []}
    db.close()

    out_dir = files[0].parent
    date = datetime.now(timezone.utc).strftime("%Y%m%d")
    if not args.dry_run:
        base = _write_reports(out_dir, date, agree)
        print(f"합의: items={agree['n_items']} {agree['agreement_breakdown']}  κ={ {k: v['mean_kappa'] for k, v in agree['kappa'].items()} }")
        print(f"리포트 → {base.relative_to(PROJECT_ROOT)}.{{json,md,csv}}  (disagreements {len(agree['disagreements'])})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
