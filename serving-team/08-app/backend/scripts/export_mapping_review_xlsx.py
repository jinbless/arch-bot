#!/usr/bin/env python3
"""HITL: ohs_analysis_records → reviewer별 검토용 .xlsx (SR/Guide/Scene 3시트).

읽기 전용 — PG/서빙/매핑/status 미변경. 검토 결과는 import_mapping_review_xlsx.py로 다시 적재.
provenance 컬럼은 잠금, verdict 컬럼만 편집 가능 + 드롭다운. 썸네일(data-URI)은 셀에 임베드.

사용:
  PYTHONIOENCODING=utf-8 python scripts/export_mapping_review_xlsx.py --reviewers alice,bob [--limit 50] [--image-only]
"""
from __future__ import annotations

import argparse
import base64
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[4]  # scripts→backend→08-app→serving-team→arch-bot
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # _mapping_review_common

from openpyxl import Workbook  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill, Protection  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from app.db.database import SessionLocal  # noqa: E402
from app.db.models import OhsAnalysisRecord  # noqa: E402
import _mapping_review_common as mc  # noqa: E402

DEFAULT_OUT = PROJECT_ROOT / "data-team" / "05-enrichment" / "eval-data" / "review-xlsx"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
UNLOCK_FILL = PatternFill("solid", fgColor="FFF2CC")  # 편집칸 시각 강조
IMG_W = 150  # px


def _thumbnail(data_uri: str):
    """data:image/...;base64,XXXX → (XLImage, row_height_px) | None."""
    if not data_uri or "base64," not in data_uri:
        return None
    try:
        from PIL import Image as PILImage
        raw = base64.b64decode(data_uri.split("base64,", 1)[1])
        pil = PILImage.open(BytesIO(raw))
        w, h = pil.size
        scale = IMG_W / float(w)
        out = BytesIO()
        pil.convert("RGB").save(out, format="PNG")
        out.seek(0)
        xl = XLImage(out)
        xl.width = IMG_W
        xl.height = int(h * scale)
        return xl, xl.height
    except Exception:  # noqa: BLE001
        return None


def _add_sheet(wb: Workbook, title: str, cols: list, rows: list[dict], reviewer: str):
    ws = wb.create_sheet(title)
    headers = [c[1] for c in cols] + ["reviewer"]
    ws.append(headers)
    for ci, _h in enumerate(headers, 1):
        cell = ws.cell(1, ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 컬럼 너비 + DV 등록
    dv_map = {}
    photo_idx = None
    for ci, (key, _hdr, width, _unlocked, dd) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        if key == "photo":
            photo_idx = ci
        if dd:
            dv_map.setdefault(dd, []).append(ci)
    ws.column_dimensions[get_column_letter(len(headers))].width = 14  # reviewer
    dvs = {
        "verdict": DataValidation(type="list", formula1='"%s"' % ",".join(mc.VERDICT_OPTIONS), allow_blank=True),
        "reason": DataValidation(type="list", formula1='"%s"' % ",".join(mc.REASON_CODES), allow_blank=True),
        "yesno": DataValidation(type="list", formula1='"%s"' % ",".join(mc.YESNO_OPTIONS), allow_blank=True),
    }
    for dv in dvs.values():
        ws.add_data_validation(dv)

    for r, row in enumerate(rows, start=2):
        for ci, (key, _hdr, _w, unlocked, _dd) in enumerate(cols, 1):
            cell = ws.cell(r, ci)
            if key == "photo":
                emb = _thumbnail(row.get("photo") or "")
                if emb:
                    xl, hpx = emb
                    ws.add_image(xl, f"{get_column_letter(ci)}{r}")
                    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, hpx * 0.75)
                    cell.value = None
                else:
                    cell.value = "(이미지 없음)"
            else:
                cell.value = row.get(key, "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if unlocked:
                cell.protection = Protection(locked=False)
                cell.fill = UNLOCK_FILL
        # reviewer 컬럼: pre-fill + locked
        rc = ws.cell(r, len(headers))
        rc.value = reviewer
    # DV 영역 부여(데이터 끝까지)
    last = len(rows) + 1
    for dd, idxs in dv_map.items():
        for ci in idxs:
            col = get_column_letter(ci)
            dvs[dd].add(f"{col}2:{col}{max(last, 2)}")
    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    return ws


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--reviewers", default="reviewer1", help="콤마구분 reviewer id (파일별 사본)")
    ap.add_argument("--limit", type=int, default=0, help="레코드 수 제한(0=전체)")
    ap.add_argument("--image-only", action="store_true", help="analysis_type='image'만")
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    db = SessionLocal()
    q = db.query(OhsAnalysisRecord).order_by(OhsAnalysisRecord.created_at)
    if args.image_only:
        q = q.filter(OhsAnalysisRecord.analysis_type == "image")
    if args.limit:
        q = q.limit(args.limit)
    records = q.all()

    sheets: dict[str, list[dict]] = {"SR_mapping": [], "GUIDE_mapping": [], "SCENE": []}
    for rec in records:
        built = mc.build_rows_for_record(db, rec.id, rec.result_json or {}, rec.image_path)
        for k in sheets:
            sheets[k].extend(built[k])
    db.close()

    n = {k: len(v) for k, v in sheets.items()}
    print(f"records={len(records)}  rows: SR={n['SR_mapping']} GUIDE={n['GUIDE_mapping']} SCENE={n['SCENE']}")

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    date = datetime.now(timezone.utc).strftime("%Y%m%d")
    reviewers = [r.strip() for r in args.reviewers.split(",") if r.strip()]
    written = []
    for reviewer in reviewers:
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        for title, (_kind, cols) in mc.SHEETS.items():
            _add_sheet(wb, title, cols, sheets[title], reviewer)
        fpath = out_dir / f"ohs_mapping_review_{date}__{reviewer}.xlsx"
        wb.save(str(fpath))
        written.append(fpath)
        print(f"  wrote {fpath.relative_to(PROJECT_ROOT)}")
    print(f"DONE — {len(written)} reviewer 파일")
    return 0


if __name__ == "__main__":
    sys.exit(main())
