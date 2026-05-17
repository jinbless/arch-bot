#!/usr/bin/env python3
"""Parse KOSHA official reference files into structured JSON for Phase 3A prompts.

Inputs (data-team/05-enrichment/reference-data/):
- 한국산업안전보건공단_산업중분류별_발생형태별_사고재해자수_*.csv  → KOSHA 22대 사고유형
- 한국표준산업분류(11차)_분류항목표.xlsx                          → KSIC 11차
- 1_전문_화학물질_및_물리적_인자의_노출기준.xlsx                  → OEL 화학물질
- 2026년도_산재보험_사업종류별_보험료율.xlsx                       → 산재보험 사업종류 (KSIC 대응)

Output:
- runtime-artifacts/kosha_reference_parsed.json
  {
    "accident_types_22": [{"ko": "떨어짐", "en_suggested": "FALL", ...}, ...],
    "ksic_categories": [{"code": "A", "name_ko": "농업·임업·어업", ...}, ...],
    "oel_chemicals": [{"name_ko": "...", "cas": "...", "oel_twa": ...}, ...],
    "insurance_business_types": [{"code": "...", "name_ko": "...", "rate": ...}, ...]
  }
"""
from __future__ import annotations
import csv
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
REF_DIR = ROOT / "data-team/05-enrichment/reference-data"
OUT_PATH = ROOT / "data-team/05-enrichment/runtime-artifacts/kosha_reference_parsed.json"

CSV_FILE = REF_DIR / "한국산업안전보건공단_산업중분류별_발생형태별_사고재해자수_20241231.csv"
KSIC_FILE = REF_DIR / "한국표준산업분류(11차)_분류항목표.xlsx"
OEL_FILE = REF_DIR / "1_전문_화학물질_및_물리적_인자의_노출기준.xlsx"
INSURANCE_FILE = REF_DIR / "2026년도_산재보험_사업종류별_보험료율.xlsx"


# KOSHA 22대 분류의 권장 EN 매핑 (한국어 라벨 + 권장 enum code)
# Phase 3A LLM prompt에 ground truth로 주입
KOSHA_22_EN_SUGGESTED = {
    "떨어짐": ("FALL", "추락 (고소/저공)"),
    "넘어짐": ("SLIP_TRIP", "미끄러짐·넘어짐"),
    "부딪힘": ("COLLISION", "부딪힘·충돌"),
    "맞음": ("STRUCK_BY", "비래·낙하물에 맞음"),
    "무너짐": ("COLLAPSE", "구조물 붕괴"),
    "끼임": ("CAUGHT_IN", "끼임·협착 (CRUSH보다 광의)"),
    "절단베임찔림": ("CUT_LACERATION", "절단·베임·찔림"),
    "감전": ("ELECTRIC_SHOCK", "감전"),
    "폭발파열": ("EXPLOSION", "폭발·파열"),
    "화재": ("FIRE_INJURY", "화재 (화상과 구분 — 화재로 인한 부상)"),
    "깔림뒤집힘": ("CRUSHED_OVERTURNED", "깔림·뒤집힘"),
    "이상온도물체접촉": ("TEMP_EXTREME_CONTACT", "이상온도 물체 접촉 (화상/동상)"),
    "빠짐익사": ("DROWNING", "빠짐·익사"),
    "불균형및무리한동작": ("ERGONOMIC_STRAIN", "불균형 및 무리한 동작 (현 catalog ERGONOMIC)"),
    "화학물질누출접촉": ("CHEMICAL_EXPOSURE", "화학물질 누출·접촉"),
    "산소결핍": ("OXYGEN_DEFICIENCY", "산소결핍 질식"),
    "사업장내교통사고": ("WORKPLACE_TRAFFIC", "사업장 내 교통사고"),
    "사업장외교통사고": ("OFF_SITE_TRAFFIC", "사업장 외 교통사고"),
    "체육행사": ("SPORTS_EVENT_INJURY", "체육행사 중 부상"),
    "폭력행위": ("VIOLENCE", "폭력행위"),
    "동물상해": ("ANIMAL_INJURY", "동물상해"),
    "기타": ("OTHER_ACCIDENT", "기타"),
    "분류불능": ("UNCLASSIFIED", "분류불능"),
}


def parse_accident_types_22() -> list[dict]:
    """CSV 컬럼에서 22대 사고유형 추출 + 산업×사고 통계 집계."""
    if not CSV_FILE.exists():
        return []
    with CSV_FILE.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        # 첫 4개 컬럼 = [대업종, 중업종, 규모, 총계], 그 뒤가 사고유형들
        accident_cols = header[4:]
        # 집계: 각 사고유형의 industry × counts
        totals: dict[str, dict] = {ko: {"total_count": 0, "by_industry": {}} for ko in accident_cols}
        for row in reader:
            if len(row) < len(header):
                continue
            대업종 = row[0]
            counts = row[4:]
            for ko, c in zip(accident_cols, counts):
                try:
                    n = int(c)
                except (ValueError, TypeError):
                    n = 0
                if n > 0:
                    totals[ko]["total_count"] += n
                    totals[ko]["by_industry"][대업종] = totals[ko]["by_industry"].get(대업종, 0) + n
    out = []
    for ko in accident_cols:
        en_suggested, description = KOSHA_22_EN_SUGGESTED.get(ko, ("", ""))
        # top 5 industries
        sorted_inds = sorted(totals[ko]["by_industry"].items(), key=lambda x: -x[1])[:5]
        out.append({
            "ko": ko,
            "en_suggested": en_suggested,
            "description_ko": description,
            "total_count_2024": totals[ko]["total_count"],
            "top_5_industries": [{"industry": ind, "count": n} for ind, n in sorted_inds],
        })
    return out


def parse_ksic_xlsx() -> list[dict]:
    """KSIC 11차 분류 항목표 파싱."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not KSIC_FILE.exists():
        return []
    wb = load_workbook(KSIC_FILE, read_only=True, data_only=True)
    ws = wb.active
    out = []
    seen_codes = set()
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if headers is None and row and any("분류" in str(c) or "코드" in str(c) for c in row if c):
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if headers is None:
            continue
        # We're looking for sections. Pull all non-empty cells, classify by depth (col 0 = 대분류, ...)
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        # Heuristic: first non-empty cell = code (letter or number)
        for j, val in enumerate(cells):
            if val and j < 5:
                # 코드 패턴 확인
                if val.replace(".", "").isalnum() and len(val) <= 6:
                    if val in seen_codes:
                        break
                    seen_codes.add(val)
                    # name in next column
                    name = cells[j+1] if j+1 < len(cells) else ""
                    out.append({"code": val, "name_ko": name, "level": j})
                break
    return out[:200]  # KSIC has many - limit for prompt size


def parse_oel_xlsx() -> list[dict]:
    """OEL 화학물질 노출기준 파싱."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not OEL_FILE.exists():
        return []
    wb = load_workbook(OEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    out = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if headers is None and any("물질" in c or "명칭" in c or "CAS" in c for c in cells):
            headers = cells
            continue
        if headers is None:
            continue
        # row data
        entry = {}
        for h, v in zip(headers, cells):
            if h:
                entry[h] = v
        # Filter: must have substance name
        name_keys = [k for k in entry if "물질" in k or "명칭" in k]
        if not name_keys or not entry.get(name_keys[0]):
            continue
        out.append(entry)
    return out[:300]  # limit


def parse_insurance_xlsx() -> list[dict]:
    """산재보험 사업종류별 보험료율 파싱."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not INSURANCE_FILE.exists():
        return []
    wb = load_workbook(INSURANCE_FILE, read_only=True, data_only=True)
    ws = wb.active
    out = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if headers is None and any("사업" in c or "업종" in c for c in cells):
            headers = cells
            continue
        if headers is None:
            continue
        entry = {h: v for h, v in zip(headers, cells) if h}
        if entry:
            out.append(entry)
    return out


def main():
    print(f"Parsing KOSHA reference from {REF_DIR}")
    data = {
        "generated_at": "2026-05-17",
        "source_files": {
            "accident_22": CSV_FILE.name,
            "ksic_11": KSIC_FILE.name,
            "oel": OEL_FILE.name,
            "insurance": INSURANCE_FILE.name,
        },
        "accident_types_22": parse_accident_types_22(),
        "ksic_categories": parse_ksic_xlsx(),
        "oel_chemicals": parse_oel_xlsx(),
        "insurance_business_types": parse_insurance_xlsx(),
    }
    print(f"  accident_types_22 : {len(data['accident_types_22'])}")
    print(f"  ksic_categories   : {len(data['ksic_categories'])}")
    print(f"  oel_chemicals     : {len(data['oel_chemicals'])}")
    print(f"  insurance         : {len(data['insurance_business_types'])}")
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSaved: {OUT_PATH}")


if __name__ == "__main__":
    main()
